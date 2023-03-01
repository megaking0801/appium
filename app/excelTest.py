import csv
from appium import webdriver
from appium.webdriver.common.appiumby import AppiumBy
import datetime
import openpyxl



wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
wb.save('empty.xlsx')
wb = openpyxl.load_workbook('empty.xlsx', data_only=True) #讀取excel

#防止excel資料表被洗掉
#dt = datetime.datetime.now()
#excelNameTime = str(dt.year) +'年'+str(dt.month)+'月'+str(dt.day)+'日'+str(dt.hour)+'點'+str(dt.minute)+'分'
#wb = openpyxl.Workbook()    # 建立空白的 Excel 活頁簿物件
#wb.save('empty.xlsx'+excelNameTime)
#wb = openpyxl.load_workbook('empty.xlsx'+excelNameTime, data_only=True) #讀取excel    

class PrintExcel:
    #印標題
    def title(driver):
        data = [["測試內容","測試結果","備註","測試時間"]]
        sheet=wb['Sheet']       
        for i in data:
            sheet.append(i)                # 逐筆添加到最後一列
        wb.save('empty.xlsx')

    #檢測測試時間
    def testTime(driver):
        dt = datetime.datetime.now()
        time = str(dt.year) +'年'+str(dt.month)+'月'+str(dt.day)+'日'+str(dt.hour)+'點'+str(dt.minute)+'分'
        return time

    #測試公司
    def testCompany(driver,testName ,result):
        data = [[testName ,result ,"" ,PrintExcel.testTime(driver)]]
        #要寫入的資料表   
        sheet=wb['Sheet']
        for i in data:
            sheet.append(i)                   # 逐筆添加到最後一列

        wb.save('empty.xlsx')

    #印出零股測試結果
    def testOddLot(driver,companyName,result):
        data = [[companyName + " 零股名稱測試結果" ,result , "" ,PrintExcel.testTime(driver)]]
        #要寫入的資料表   
        sheet=wb['Sheet']
        for i in data:
            sheet.append(i)                   # 逐筆添加到最後一列

        wb.save('empty.xlsx')

    #印出K線測試結果
    def testKLine(driver,companyName,result):
        data = [[companyName + " K線名稱測試結果" ,result , "" ,PrintExcel.testTime(driver)]]
        #要寫入的資料表   
        sheet=wb['Sheet']
        for i in data:
            sheet.append(i)                   # 逐筆添加到最後一列

        wb.save('empty.xlsx')

    #印出外資、投信...等測試資料
    def testNumber(driver,testNum,result,company,compareName):
        data = [[ company + " " + compareName + " " + testNum ,result , "" ,PrintExcel.testTime(driver) ]]
        #要寫入的資料表 
        sheet=wb['Sheet'] 
        for i in data:
            sheet.append(i)                   # 逐筆添加到最後一列

        wb.save('empty.xlsx')   

    #印出買入、賣出測試結果
    def clickBuyAndSell(driver,companyName,testName,result):
        
        data = [[ companyName + testName ,result , "" ,PrintExcel.testTime(driver) ]]
        #要寫入的資料表 
        sheet=wb['Sheet'] 
        for i in data:
            sheet.append(i)                   # 逐筆添加到最後一列

        wb.save('empty.xlsx')     

    def error(driver):
        data = [[ "無法找到資料或發生錯誤" ,PrintExcel.testTime(driver) ]]
        #要寫入的資料表 
        sheet=wb['Sheet'] 
        for i in data:
            sheet.append(i)                   # 逐筆添加到最後一列

        wb.save('empty.xlsx')   

             

    
	   	   